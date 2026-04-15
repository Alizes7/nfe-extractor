import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List
import io
from datetime import datetime

class NFeExcelGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Estilos
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.subheader_font = Font(name='Calibri', size=10, bold=True, color="000000")
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        self.money_format = '#,##0.00'
        self.date_format = 'DD/MM/YYYY'
        
    def _ajustar_largura_colunas(self, ws):
        """Ajusta largura das colunas automaticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _aplicar_estilo_header(self, cell):
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self.border
    
    def _aplicar_estilo_subheader(self, cell):
        cell.fill = self.subheader_fill
        cell.font = self.subheader_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.border
    
    def _aplicar_estilo_dado(self, cell, is_number=False, is_money=False):
        cell.border = self.border
        cell.alignment = Alignment(horizontal='right' if is_number else 'left', vertical='center')
        if is_money:
            cell.number_format = self.money_format
        elif is_number:
            cell.number_format = '#,##0.00'
    
    def criar_aba_resumo(self, dados_list: List[Dict]):
        """Cria aba de resumo com todas as notas"""
        ws = self.wb.create_sheet("Resumo NF-es")
        
        # Headers
        headers = [
            'Arquivo', 'Número NF', 'Série', 'Data Emissão', 'Tipo', 
            'Emitente', 'CNPJ Emitente', 'Destinatário', 'CNPJ Destinatário',
            'Valor Produtos', 'Valor ICMS', 'Desconto', 'Valor Total', 
            'Qtd Itens', 'Chave Acesso'
        ]
        
        ws.append(headers)
        for cell in ws[1]:
            self._aplicar_estilo_header(cell)
        
        # Dados
        for dado in dados_list:
            cab = dado['cabecalho']
            emit = dado['emitente']
            dest = dado['destinatario']
            tot = dado['totais']
            
            row = [
                dado['nome_arquivo'],
                cab['numero'],
                cab['serie'],
                cab['data_emissao'],
                cab['tipo_operacao'],
                emit['nome'],
                emit['cnpj'],
                dest['nome'],
                dest['cnpj'],
                tot['valor_produtos'],
                tot['valor_icms'],
                tot['desconto'],
                tot['valor_total'],
                dado['quantidade_produtos'],
                cab['chave_acesso']
            ]
            ws.append(row)
            
            # Aplicar formatação na linha
            for idx, cell in enumerate(ws[ws.max_row], start=1):
                is_money = idx in [10, 11, 12, 13]  # Colunas de valores
                self._aplicar_estilo_dado(cell, is_number=is_money, is_money=is_money)
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Filtro
        ws.auto_filter.ref = ws.dimensions
        
        self._ajustar_largura_colunas(ws)
    
    def criar_aba_detalhada(self, dados: Dict, nome_aba: str = None):
        """Cria aba detalhada para uma NF-e específica"""
        num_nf = dados['cabecalho']['numero']
        nome_aba = nome_aba or f"NF-{num_nf}"
        # Limita nome da aba em 31 caracteres (limite do Excel)
        nome_aba = nome_aba[:31]
        
        ws = self.wb.create_sheet(nome_aba)
        
        # TÍTULO
        ws.merge_cells('A1:H1')
        ws['A1'] = f"NOTA FISCAL ELETRÔNICA - Nº {num_nf}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # CABEÇALHO
        ws.merge_cells('A3:H3')
        ws['A3'] = "DADOS DA NOTA FISCAL"
        ws['A3'].fill = self.header_fill
        ws['A3'].font = self.header_font
        ws['A3'].alignment = Alignment(horizontal='left')
        
        row = 4
        cab = dados['cabecalho']
        ws.append(['Número:', cab['numero'], 'Série:', cab['serie'], 
                   'Data Emissão:', cab['data_emissao'], 'Tipo:', cab['tipo_operacao']])
        ws.append(['Natureza:', cab['natureza_operacao'], '', '', '', '', '', ''])
        ws.append(['Chave Acesso:', cab['chave_acesso'], '', '', '', '', '', ''])
        ws.append(['Protocolo:', cab['protocolo'], '', '', '', '', '', ''])
        
        for r in range(4, 8):
            for cell in ws[r]:
                self._aplicar_estilo_subheader(cell) if cell.column <= 2 else None
        
        row = 9
        
        # EMITENTE
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "EMITENTE"
        ws[f'A{row}'].fill = self.header_fill
        ws[f'A{row}'].font = self.header_font
        row += 1
        
        emit = dados['emitente']
        ws.append(['Razão Social:', emit['nome'], 'CNPJ:', emit['cnpj'], 
                   'IE:', emit['ie'], 'Telefone:', emit['telefone']])
        ws.append(['Endereço:', emit['endereco'], 'Bairro:', emit['bairro'],
                   'CEP:', emit['cep'], 'Município:', f"{emit['municipio']}-{emit['uf']}"])
        row += 3
        
        # DESTINATÁRIO
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "DESTINATÁRIO"
        ws[f'A{row}'].fill = self.header_fill
        ws[f'A{row}'].font = self.header_font
        row += 1
        
        dest = dados['destinatario']
        ws.append(['Razão Social:', dest['nome'], 'CNPJ:', dest['cnpj'], 
                   'IE:', dest['ie'], 'Telefone:', dest['telefone']])
        ws.append(['Endereço:', dest['endereco'], 'Bairro:', dest['bairro'],
                   'CEP:', dest['cep'], 'Município:', f"{dest['municipio']}-{dest['uf']}"])
        row += 3
        
        # PRODUTOS
        ws.merge_cells(f'A{row}:O{row}')
        ws[f'A{row}'] = "PRODUTOS E SERVIÇOS"
        ws[f'A{row}'].fill = self.header_fill
        ws[f'A{row}'].font = self.header_font
        row += 1
        
        headers_prod = ['Código', 'Descrição', 'NCM', 'CST', 'CFOP', 'UN', 
                       'Qtd', 'V. Unit', 'V. Total', 'Desc', 'BC ICMS', 
                       'V. ICMS', 'Aliq ICMS', 'V. IPI', 'Aliq IPI']
        ws.append(headers_prod)
        for cell in ws[row]:
            self._aplicar_estilo_header(cell)
        row += 1
        
        # Dados dos produtos
        for prod in dados['produtos']:
            ws.append([
                prod['codigo'],
                prod['descricao'],
                prod['ncm'],
                prod['cst'],
                prod['cfop'],
                prod['unidade'],
                prod['quantidade'],
                prod['valor_unitario'],
                prod['valor_total'],
                prod['desconto'],
                prod['base_icms'],
                prod['valor_icms'],
                prod['aliq_icms'],
                prod['valor_ipi'],
                prod['aliq_ipi']
            ])
            
            # Formatação
            for idx, cell in enumerate(ws[ws.max_row], start=1):
                is_num = idx >= 7  # Colunas numéricas a partir da 7ª
                is_money = idx in [8, 9, 10, 11, 12, 13, 14]  # Valores monetários
                self._aplicar_estilo_dado(cell, is_number=is_num, is_money=is_money)
        
        row = ws.max_row + 2
        
        # TOTAIS
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "TOTAIS"
        ws[f'A{row}'].fill = self.header_fill
        ws[f'A{row}'].font = self.header_font
        row += 1
        
        tot = dados['totais']
        totais_data = [
            ['Base Cálculo ICMS:', tot['base_icms'], 'Valor ICMS:', tot['valor_icms']],
            ['Base Cálculo ST:', tot['base_icms_st'], 'Valor ICMS ST:', tot['valor_icms_st']],
            ['Valor Produtos:', tot['valor_produtos'], 'Valor Frete:', tot['valor_frete']],
            ['Valor Seguro:', tot['valor_seguro'], 'Desconto:', tot['desconto']],
            ['Outras Despesas:', tot['outras_despesas'], 'Valor Aprox. Tributos:', tot['valor_aprox_tributos']],
            ['', '', 'VALOR TOTAL DA NOTA:', tot['valor_total']]
        ]
        
        for linha in totais_data:
            ws.append(linha)
            for idx, cell in enumerate(ws[ws.max_row], start=1):
                if idx in [2, 4]:  # Colunas de valores
                    cell.number_format = self.money_format
                    cell.font = Font(bold=True) if 'VALOR TOTAL' in str(linha[0]) or 'VALOR TOTAL' in str(linha[2]) else Font()
                cell.border = self.border
        
        self._ajustar_largura_colunas(ws)
    
    def gerar_excel(self, dados_list: List[Dict]) -> bytes:
        """Gera arquivo Excel em memória"""
        # Cria aba de resumo
        self.criar_aba_resumo(dados_list)
        
        # Cria aba detalhada para cada NF
        for dados in dados_list:
            self.criar_aba_detalhada(dados)
        
        # Salva em memória
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()
